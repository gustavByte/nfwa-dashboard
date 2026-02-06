from __future__ import annotations

import argparse
import csv
import re
from datetime import datetime, time
from pathlib import Path
from typing import Iterable, Optional

import openpyxl


DEFAULT_XLSX = Path("data/manual_sources/kondis/2003 -1997 Norgesstatistikk maraton menn.xlsx")
DEFAULT_OUT_DIR = Path("nfwa/reference_data")
OUT_NAME_TEMPLATE = "kondis_{season}_maraton_menn.csv"


def _none_if_empty(value: object) -> Optional[str]:
    text = str(value or "").strip()
    return text or None


def _parse_int(value: object) -> Optional[int]:
    if value is None:
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value) if value.is_integer() else None
    text = str(value).strip()
    if not text:
        return None
    m = re.search(r"\d{1,4}", text)
    if not m:
        return None
    try:
        return int(m.group(0))
    except ValueError:
        return None


def _format_marathon_time(value: object) -> Optional[str]:
    if value is None:
        return None

    if isinstance(value, datetime):
        value = value.time()

    if isinstance(value, time):
        return f"{int(value.hour)}:{int(value.minute):02d}:{int(value.second):02d}"

    if isinstance(value, (int, float)):
        # Excel can store times as day fractions.
        day_fraction = float(value)
        if 0 <= day_fraction < 1:
            total = int(round(day_fraction * 86400))
            h = total // 3600
            m = (total % 3600) // 60
            s = total % 60
            return f"{h}:{m:02d}:{s:02d}"

    text = str(value).strip()
    if not text:
        return None

    cleaned = re.sub(r"\s+", "", text)
    cleaned = cleaned.replace(",", ".")
    cleaned = re.sub(r"[hH]$", "", cleaned)

    parts = [p for p in re.split(r"[:.]", cleaned) if p]
    if len(parts) not in {2, 3} or not all(p.isdigit() for p in parts):
        return None

    h = int(parts[0])
    m = int(parts[1])
    s = int(parts[2]) if len(parts) == 3 else 0
    if not (0 <= m < 60 and 0 <= s < 60):
        return None
    return f"{h}:{m:02d}:{s:02d}"


def _season_time_col(header_row: Iterable[object]) -> int:
    cells = list(header_row)
    for i, raw in enumerate(cells, start=1):
        low = str(raw or "").strip().lower()
        if not low:
            continue
        if "tid" in low and "pb" not in low and "pr" not in low:
            return i
    return 6


def _extract_sheet_rows(*, sheet: openpyxl.worksheet.worksheet.Worksheet, season: int) -> list[dict[str, object]]:
    time_col = _season_time_col(sheet.iter_rows(min_row=1, max_row=1, values_only=True).__next__())
    out: list[dict[str, object]] = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        rank_cell = row[0] if len(row) > 0 else None
        if isinstance(rank_cell, str) and rank_cell.strip().lower().startswith("kilde"):
            continue

        rank = _parse_int(rank_cell)
        if rank is None:
            continue

        athlete_name = _none_if_empty(row[1] if len(row) > 1 else None)
        if not athlete_name:
            continue

        raw_time = row[time_col - 1] if len(row) >= time_col else None
        performance_raw = _format_marathon_time(raw_time)
        if not performance_raw:
            continue

        out.append(
            {
                "season": season,
                "rank_in_list": rank,
                "athlete_name": athlete_name,
                "club_name": _none_if_empty(row[2] if len(row) > 2 else None) or "",
                "birth_year": _parse_int(row[3] if len(row) > 3 else None) or "",
                "venue_city": _none_if_empty(row[4] if len(row) > 4 else None) or "",
                "performance_raw": performance_raw,
            }
        )

    return out


def _sheet_season(sheet_name: str) -> Optional[int]:
    m = re.fullmatch(r"\s*(\d{4})\s*", sheet_name or "")
    if not m:
        return None
    return int(m.group(1))


def _write_csv(*, rows: list[dict[str, object]], csv_path: Path) -> None:
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    with csv_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=["rank_in_list", "athlete_name", "club_name", "birth_year", "venue_city", "performance_raw"],
        )
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {
                    "rank_in_list": row["rank_in_list"],
                    "athlete_name": row["athlete_name"],
                    "club_name": row["club_name"],
                    "birth_year": row["birth_year"],
                    "venue_city": row["venue_city"],
                    "performance_raw": row["performance_raw"],
                }
            )


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Generer manuelle Kondis-CSV-er for maraton menn fra korreksjonsarbeidsbok (ignorerer PB/PR-kolonner)."
    )
    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX, help="Kilderegnark (*.xlsx)")
    parser.add_argument("--out-dir", type=Path, default=DEFAULT_OUT_DIR, help="Utmappe for CSV-filer")
    parser.add_argument(
        "--years",
        nargs="+",
        type=int,
        default=[2003, 2002, 2001, 2000, 1999, 1998, 1997],
        help="Hvilke sesonger som skal eksporteres",
    )
    args = parser.parse_args()

    if not args.xlsx.exists():
        raise FileNotFoundError(f"Fant ikke kilderegnark: {args.xlsx}")

    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    requested_years = {int(y) for y in args.years}
    found_years: set[int] = set()
    exported = 0

    for sheet in wb.worksheets:
        season = _sheet_season(sheet.title)
        if season is None or season not in requested_years:
            continue
        found_years.add(season)

        rows = _extract_sheet_rows(sheet=sheet, season=season)
        out_path = args.out_dir / OUT_NAME_TEMPLATE.format(season=season)
        _write_csv(rows=rows, csv_path=out_path)
        exported += 1
        print(f"{season}: skrev {len(rows)} rader -> {out_path}")

    missing = sorted(requested_years - found_years)
    if missing:
        print(f"Advarsel: fant ikke ark for sesonger: {missing}")

    print(f"Ferdig. Eksporterte {exported} CSV-fil(er).")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
